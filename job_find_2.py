from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import json
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

service = Service(
    r"C:\Users\Admin\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"
)

driver = webdriver.Chrome(service=service)
driver.maximize_window()
wait = WebDriverWait(driver, 20)

email = "parikhdhruv05@gmail.com"
password = "dhruvparikh@1234"


# ============================================================
# 🆕 DATE-WISE OUTPUT FOLDER SETUP
# ============================================================

today_str = datetime.now().strftime("%Y-%m-%d")
BASE_OUTPUT_DIR = "linkedin_posts_output"
OUTPUT_DIR = os.path.join(BASE_OUTPUT_DIR, today_str)
os.makedirs(OUTPUT_DIR, exist_ok=True)
print(f"📁 Output folder ready: {OUTPUT_DIR}")


# ============================================================
# BOLD UNICODE NORMALIZATION
# ============================================================

def build_bold_map():
    bold_map = {}
    for i in range(26):
        bold_map[chr(0x1D400 + i)] = chr(ord('A') + i)
        bold_map[chr(0x1D41A + i)] = chr(ord('a') + i)
    for i in range(10):
        bold_map[chr(0x1D7CE + i)] = chr(ord('0') + i)
    return bold_map


BOLD_MAP = build_bold_map()


def normalize_bold(text):
    return "".join(BOLD_MAP.get(ch, ch) for ch in text)


# ============================================================
# EXTRACTION FUNCTIONS
# ============================================================

def extract_emails(text):
    normal_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
    normalized_text = normalize_bold(text)
    bold_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', normalized_text)
    all_emails = list(set(normal_emails + bold_emails))
    return ", ".join(all_emails) if all_emails else ""


def extract_phone_numbers(text):
    normalized_text = normalize_bold(text)
    pattern = r'(?:\+?91[\s-]?)?[6-9]\d{9}'
    matches = re.findall(pattern, normalized_text)
    spaced_pattern = r'(?:\+?91[\s-]?)?[6-9]\d{4}[\s-]?\d{5}'
    spaced_matches = re.findall(spaced_pattern, normalized_text)
    all_numbers = set()
    for m in matches + spaced_matches:
        cleaned = re.sub(r'[\s-]', '', m)
        if len(cleaned) >= 10:
            all_numbers.add(cleaned)
    return ", ".join(all_numbers) if all_numbers else ""


def extract_apply_links(text):
    """
    🆕 Extract apply / WhatsApp / form links from post text.
    Priority order:
      1. Direct apply links (lnkd.in, forms, greenhouse, lever, workday, etc.)
      2. WhatsApp links
      3. Any https link near keywords 'apply', 'join', 'form'
    Returns comma-separated unique links.
    """
    normalized = normalize_bold(text)

    # Pattern 1: lnkd.in short links (LinkedIn safety-wrapped or direct)
    lnkd_pattern = r'https?://lnkd\.in/[^\s\)\]>\"\'<]+'
    lnkd_links = re.findall(lnkd_pattern, normalized)

    # Pattern 2: Common ATS / job board URLs
    ats_pattern = (
        r'https?://(?:'
        r'(?:[\w-]+\.)?greenhouse\.io|'
        r'(?:[\w-]+\.)?lever\.co|'
        r'(?:[\w-]+\.)?workday\.com|'
        r'(?:[\w-]+\.)?naukri\.com|'
        r'(?:[\w-]+\.)?instahyre\.com|'
        r'(?:[\w-]+\.)?internshala\.com|'
        r'(?:[\w-]+\.)?unstop\.com|'
        r'(?:[\w-]+\.)?cutshort\.io|'
        r'(?:[\w-]+\.)?wellfound\.com|'
        r'(?:[\w-]+\.)?angellist\.com|'
        r'(?:[\w-]+\.)?indeed\.com|'
        r'(?:[\w-]+\.)?foundit\.in|'
        r'wa\.me|'
        r'chat\.whatsapp\.com|'
        r'forms\.gle|'
        r'docs\.google\.com/forms'
        r')[^\s\)\]>\"\'<]+'
    )
    ats_links = re.findall(ats_pattern, normalized)

    # Pattern 3: Any URL that appears near apply/join/career keywords (within 100 chars)
    keyword_context = re.findall(
        r'(?:apply|Apply|APPLY|join|Join|career|Career|form|Form|link|Link|here|Here)'
        r'.{0,80}?(https?://[^\s\)\]>\"\'<]+)',
        normalized
    )

    all_links = []
    seen = set()
    for link in lnkd_links + ats_links + keyword_context:
        # Clean trailing punctuation
        link = re.sub(r'[.,;:!?\)\]]+$', '', link)
        if link not in seen and len(link) > 10:
            seen.add(link)
            all_links.append(link)

    return ", ".join(all_links[:5]) if all_links else ""


def extract_position(text):
    patterns = [
        r'(?:Hiring for|Position|Role|Job Title|We are hiring|hiring)[:\-]?\s*([A-Za-z0-9 /&,\-\(\)]{3,60})',
        r'🔹\s*([A-Za-z0-9 /&,\-\(\)]{3,60})',
        r'(?:^|\n)([A-Za-z][A-Za-z0-9 /&,\-\(\)]{3,50}(?:Developer|Engineer|Manager|Executive|Designer|Analyst|Lead|Architect|Consultant|Intern|Specialist))',
    ]
    found_positions = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.MULTILINE)
        for m in matches:
            cleaned = m.strip()
            if cleaned and len(cleaned) < 70:
                found_positions.append(cleaned)
    seen = set()
    unique_positions = []
    for p in found_positions:
        if p not in seen:
            seen.add(p)
            unique_positions.append(p)
    return " | ".join(unique_positions[:5]) if unique_positions else ""


def extract_location(text):
    patterns = [
        r'(?:Location|📍)\s*[:\-]?\s*([A-Za-z ,]{2,40})',
        r'(?:Work From Office|Work From Home|Remote|Hybrid)[^\n]*',
    ]
    found_locations = []
    for pattern in patterns:
        matches = re.findall(pattern, text)
        for m in matches:
            cleaned = m.strip()
            if cleaned and len(cleaned) < 60:
                found_locations.append(cleaned)
    seen = set()
    unique_locations = []
    for loc in found_locations:
        if loc not in seen:
            seen.add(loc)
            unique_locations.append(loc)
    return " | ".join(unique_locations[:5]) if unique_locations else ""


def extract_experience(text):
    normalized_text = normalize_bold(text)
    patterns = [
        r'(?:Experience|Exp)[:\-]?\s*(\d+\+?\s*(?:-|–|to)?\s*\d*\+?\s*(?:Years|years|Yrs|yrs))',
        r'(\d+\+?\s*(?:-|–|to)\s*\d+\+?\s*(?:Years|years|Yrs|yrs))',
        r'(\d+\+\s*(?:Years|years|Yrs|yrs))',
        r'(\d+\s*(?:Years|years|Yrs|yrs)\s*(?:of\s*)?(?:experience|exp))',
    ]
    found = []
    for pattern in patterns:
        matches = re.findall(pattern, normalized_text, re.IGNORECASE)
        for m in matches:
            cleaned = m.strip()
            if cleaned:
                found.append(cleaned)
    seen = set()
    unique = []
    for e in found:
        if e.lower() not in seen:
            seen.add(e.lower())
            unique.append(e)
    return " | ".join(unique[:3]) if unique else ""


def extract_post_info(text):
    email_val   = extract_emails(text)
    mobile_val  = extract_phone_numbers(text)
    apply_val   = extract_apply_links(text)
    return {
        "email":       email_val,
        "mobile":      mobile_val,
        # 🆕 apply_link: always extracted; shown alongside email/mobile
        "apply_link":  apply_val,
        "position":    extract_position(text),
        "location":    extract_location(text),
        "experience":  extract_experience(text),
    }


# ============================================================
# 🆕 POSTER NAME / DATE / IMAGE EXTRACTION (fixed image logic)
# ============================================================

def get_poster_name_date_image(post_element, driver):
    """
    Returns (poster_name, post_date, image_url)
    Image URL: grabbed from feedshare / update-components-image <img> tags.
    Looks for 'feedshare-shrink' or 'feedshare-image' in src to avoid profile pics.
    """
    poster_name = ""
    post_date   = ""
    image_url   = ""

    # ---- find the parent feed-update container ----
    parent = None
    ancestor_xpaths = [
        "./ancestor::div[contains(@class,'feed-shared-update-v2')][1]",
        "./ancestor::div[contains(@data-urn,'urn:li:activity')][1]",
        "./ancestor::div[contains(@class,'d368900e')][1]",
        "./ancestor::div[@componentkey][1]",
    ]
    for xp in ancestor_xpaths:
        try:
            el = post_element.find_element(By.XPATH, xp)
            if el:
                parent = el
                break
        except Exception:
            continue

    if not parent:
        return poster_name, post_date, image_url

    # ---- Poster Name ----
    name_selectors = [
        ".update-components-actor__name",
        ".update-components-actor__title",
        "span.feed-shared-actor__name",
    ]
    for sel in name_selectors:
        try:
            el = parent.find_element(By.CSS_SELECTOR, sel)
            if el.text.strip():
                poster_name = el.text.strip().split("\n")[0]
                break
        except Exception:
            continue

    if not poster_name:
        try:
            els = parent.find_elements(
                By.XPATH,
                ".//div[contains(@aria-label,'Verified Profile')]//span"
            )
            for el in els:
                t = el.text.strip()
                if t and "•" not in t:
                    poster_name = t
                    break
        except Exception:
            pass

    if not poster_name:
        try:
            img_alt_el = parent.find_element(
                By.XPATH, ".//*[@aria-label[contains(.,\"profile\")]]"
            )
            alt = img_alt_el.get_attribute("aria-label") or ""
            m = re.search(r"View (.+?)['']s profile", alt)
            if m:
                poster_name = m.group(1).strip()
        except Exception:
            pass

    # ---- Post Date ----
    date_selectors = [
        ".update-components-actor__sub-description",
        "span.feed-shared-actor__sub-description",
        "time",
    ]
    for sel in date_selectors:
        try:
            el = parent.find_element(By.CSS_SELECTOR, sel)
            if el.text.strip():
                post_date = el.text.strip().split("\n")[0]
                break
        except Exception:
            continue

    if not post_date:
        try:
            spans = parent.find_elements(By.XPATH, ".//p//span")
            for sp in spans:
                t = sp.text.strip()
                if re.match(r'^\d+[smhdw]\s*•', t) or re.match(r'^\d+\s*(min|hr|day|week|month)s?\s*•', t, re.IGNORECASE):
                    post_date = t.split("•")[0].strip()
                    break
        except Exception:
            pass

    # ============================================================
    # 🆕 POST IMAGE — grab from feedshare-shrink / feedshare-image src
    #    These are actual post images, NOT profile pictures.
    #    HTML example from your paste:
    #      src="https://media.licdn.com/dms/image/v2/D5622AQG.../feedshare-shrink_160/..."
    #      src="https://media.licdn.com/dms/image/v2/D4D22AQH.../feedshare-shrink_800/..."
    # ============================================================
    try:
        all_imgs = parent.find_elements(By.TAG_NAME, "img")
        for img in all_imgs:
            src = img.get_attribute("src") or ""
            # Only pick images that are post images (feedshare-*), not profile/logo images
            if "media.licdn.com" in src and (
                "feedshare-shrink" in src
                or "feedshare-image" in src
                or "feedshare-shrink_800" in src
                or "feedshare-shrink_1280" in src
            ):
                # Prefer the highest resolution available via srcset
                srcset = img.get_attribute("srcset") or ""
                if srcset:
                    # Parse srcset: "url 20w, url 160w, url 800w" — pick largest w
                    srcset_parts = [s.strip() for s in srcset.split(",") if s.strip()]
                    best_url = src
                    best_w = 0
                    for part in srcset_parts:
                        tokens = part.split()
                        if len(tokens) == 2:
                            candidate_url = tokens[0]
                            try:
                                w = int(tokens[1].replace("w", ""))
                                if w > best_w:
                                    best_w = w
                                    best_url = candidate_url
                            except ValueError:
                                pass
                    image_url = best_url
                else:
                    image_url = src
                break  # Take the first post image found
    except Exception as e:
        print(f"  ⚠️ Image extraction error: {e}")

    return poster_name, post_date, image_url


# ============================================================
# SEE MORE / SCROLL FUNCTIONS
# ============================================================

def click_all_see_more_buttons(driver):
    clicked = 0
    try:
        see_more_selectors = [
            "button[data-testid='expandable-text-button']",
            "button.feed-shared-inline-show-more-text__see-more-less-toggle",
            "span.feed-shared-inline-show-more-text__see-more-less-toggle",
            ".see-more",
            "button[aria-label='see more']",
        ]
        for selector in see_more_selectors:
            buttons = driver.find_elements(By.CSS_SELECTOR, selector)
            for btn in buttons:
                try:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                        time.sleep(0.3)
                        driver.execute_script("arguments[0].click();", btn)
                        clicked += 1
                        time.sleep(0.2)
                except Exception:
                    continue

        xpath_buttons = driver.find_elements(
            By.XPATH,
            "//button[contains(@class,'expandable-text-button')]"
            " | //span[contains(text(),'more')][@role='button']"
            " | //button[@data-testid='expandable-text-button']"
        )
        for btn in xpath_buttons:
            try:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", btn)
                    clicked += 1
                    time.sleep(0.2)
            except Exception:
                continue
    except Exception as e:
        print(f"  ⚠️ See more click error: {e}")

    if clicked > 0:
        print(f"  🔓 {clicked} 'see more' buttons clicked")
    return clicked


def get_full_post_text(post_element, driver):
    try:
        see_more_btns = post_element.find_elements(
            By.CSS_SELECTOR, "button[data-testid='expandable-text-button']"
        )
        for btn in see_more_btns:
            try:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.5)
            except Exception:
                pass

        text = post_element.text.strip()
        if text.endswith("… more"):
            text = text[:-6].strip()
        if text.endswith("…more"):
            text = text[:-5].strip()
        return text
    except Exception:
        return ""


def scroll_and_collect_posts(driver, max_posts=100, max_no_change=5):
    all_posts = []
    all_posts_data = []
    seen_texts = set()
    no_change_count = 0
    last_post_count = 0

    print("🔍 Starting post collection with proper scroll...\n")

    while True:
        print("  🔄 Expanding all truncated posts...")
        click_all_see_more_buttons(driver)
        time.sleep(1)

        post_elements = driver.find_elements(
            By.CSS_SELECTOR, "[data-testid='expandable-text-box']"
        )

        for post in post_elements:
            try:
                full_text = get_full_post_text(post, driver)
                if full_text and full_text not in seen_texts and len(full_text) > 30:
                    seen_texts.add(full_text)
                    all_posts.append(full_text)

                    poster_name, post_date, image_url = get_poster_name_date_image(post, driver)
                    info = extract_post_info(full_text)

                    post_data = {
                        "post_number":  len(all_posts),
                        "poster_name":  poster_name,
                        "post_date":    post_date,
                        "position":     info["position"],
                        "experience":   info["experience"],
                        "location":     info["location"],
                        "email":        info["email"],
                        "mobile":       info["mobile"],
                        "apply_link":   info["apply_link"],   # 🆕
                        "image_url":    image_url,
                        "full_text":    full_text,
                    }
                    all_posts_data.append(post_data)

                    print(f"  ✅ Post #{len(all_posts)}: {full_text[:80]}...")
                    print(
                        f"     👤 {poster_name} | 🕒 {post_date} | 💼 {info['experience']} "
                        f"| 📧 {info['email']} | 📱 {info['mobile']} "
                        f"| 🔗 {info['apply_link'][:60] if info['apply_link'] else '-'} "
                        f"| 🖼️ {'Yes' if image_url else 'No'}"
                    )
            except Exception:
                continue

        print(f"\n📊 Total posts collected: {len(all_posts)}")

        if len(all_posts) >= max_posts:
            print(f"✅ Reached max posts limit: {max_posts}")
            break

        if len(all_posts) == last_post_count:
            no_change_count += 1
            print(f"⚠️  No new posts (attempt {no_change_count}/{max_no_change})")
            if no_change_count >= max_no_change:
                print("✅ No more posts to load. Stopping.")
                break
        else:
            no_change_count = 0

        last_post_count = len(all_posts)

        if post_elements:
            try:
                last_post = post_elements[-1]
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'smooth', block:'center'});", last_post
                )
                time.sleep(2)
            except Exception:
                pass

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
        driver.execute_script("window.scrollBy(0, -400);")
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)
        print(f"📜 Scrolled | Loading new posts...")

    return all_posts, all_posts_data


# ============================================================
# 🆕 EXCEL EXPORT — added Apply Link column
# ============================================================

def save_to_excel(all_posts_data, filename="linkedin_posts.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Posts"

    headers = [
        "Post #", "Poster Name", "Post Date", "Position",
        "Experience", "Location", "Email", "Mobile Number",
        "Apply Link",          # 🆕 column 9
        "Image URL",           # column 10
        "Full Post Text"       # column 11
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for data in all_posts_data:
        row = [
            data["post_number"],
            data["poster_name"],
            data["post_date"],
            data["position"],
            data["experience"],
            data["location"],
            data["email"],
            data["mobile"],
            data["apply_link"],   # 🆕 col 9
            data["image_url"],    # col 10
            data["full_text"],    # col 11
        ]
        ws.append(row)

        current_row = ws.max_row

        # 🆕 Apply Link — make clickable hyperlink (first link only)
        apply_val = data["apply_link"]
        if apply_val:
            first_link = apply_val.split(",")[0].strip()
            cell = ws.cell(row=current_row, column=9)
            cell.hyperlink = first_link
            cell.style = "Hyperlink"

        # Image URL — make clickable hyperlink
        if data["image_url"]:
            cell = ws.cell(row=current_row, column=10)
            cell.hyperlink = data["image_url"]
            cell.style = "Hyperlink"

    # 🆕 11 columns now
    col_widths = [8, 25, 15, 35, 15, 25, 35, 20, 60, 50, 80]
    for i, width in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(filename)
    print(f"📊 Excel file saved: {filename}")


# ============================================================
# MAIN SCRIPT
# ============================================================

try:
    driver.get("https://www.linkedin.com/jobs/")

    email_input = wait.until(EC.presence_of_element_located((By.ID, "session_key")))
    email_input.send_keys(email)

    password_input = wait.until(EC.presence_of_element_located((By.ID, "session_password")))
    password_input.send_keys(password)

    sign_in_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    sign_in_btn.click()
    print("✅ Login successful")
    time.sleep(5)

    job_search = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[data-testid='typeahead-input']"))
    )
    job_search.clear()
    job_search.send_keys("python developer 3 years experience and Ahmedabad")
    job_search.send_keys(Keys.ENTER)
    print("✅ Job search executed")
    time.sleep(5)

    jobs_dropdown = wait.until(
        EC.presence_of_element_located((By.XPATH, "//label[contains(.,'Jobs')]"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", jobs_dropdown)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", jobs_dropdown)
    print("✅ Jobs dropdown opened")
    time.sleep(2)

    posts_option = wait.until(
        EC.presence_of_element_located((By.XPATH, "//p[normalize-space()='Posts']"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", posts_option)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", posts_option)
    print("✅ Posts selected")
    time.sleep(2)

    sort_by = wait.until(
        EC.presence_of_element_located((By.XPATH, "//label[contains(.,'Sort by')]"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", sort_by)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", sort_by)
    print("✅ Sort By opened")
    time.sleep(2)

    latest_option = wait.until(
        EC.presence_of_element_located((By.XPATH, "//span[normalize-space()='Latest']"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", latest_option)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", latest_option)
    print("✅ Latest selected")
    time.sleep(2)

    show_results = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Show results']"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_results)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", show_results)
    print("✅ Show Results clicked")

    wait.until(EC.url_contains("/search/results/content/"))
    print("✅ Content results page loaded")
    print("Current URL:", driver.current_url)
    time.sleep(4)

    all_posts, all_posts_data = scroll_and_collect_posts(driver, max_posts=100, max_no_change=5)

    # Build full file paths inside the date-wise folder
    timestamp_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    txt_path  = os.path.join(OUTPUT_DIR, f"linkedin_posts_{timestamp_str}.txt")
    json_path = os.path.join(OUTPUT_DIR, f"linkedin_posts_{timestamp_str}.json")
    xlsx_path = os.path.join(OUTPUT_DIR, f"linkedin_posts_{timestamp_str}.xlsx")

    # ---- TXT ----
    with open(txt_path, "w", encoding="utf-8") as f:
        for i, data in enumerate(all_posts_data, 1):
            f.write(f"{'=' * 60}\n")
            f.write(f"POST #{i}\n")
            f.write(f"{'=' * 60}\n")
            f.write(f"Poster  : {data['poster_name']}\n")
            f.write(f"Date    : {data['post_date']}\n")
            f.write(f"Position: {data['position']}\n")
            f.write(f"Exp     : {data['experience']}\n")
            f.write(f"Location: {data['location']}\n")
            f.write(f"Email   : {data['email']}\n")
            f.write(f"Mobile  : {data['mobile']}\n")
            f.write(f"Apply   : {data['apply_link']}\n")   # 🆕
            f.write(f"Image   : {data['image_url']}\n")
            f.write(f"\n{data['full_text']}\n\n")

    # ---- JSON ----
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_posts_data, f, ensure_ascii=False, indent=2)

    # ---- XLSX ----
    save_to_excel(all_posts_data, xlsx_path)

    print(f"\n🎉 Done! Total {len(all_posts)} posts saved.")
    print(f"📁 Folder : {OUTPUT_DIR}")
    print(f"📄 TXT    : {txt_path}")
    print(f"📄 JSON   : {json_path}")
    print(f"📊 XLSX   : {xlsx_path}")

    time.sleep(3)

except Exception as e:
    print("❌ Error:", e)
    import traceback
    traceback.print_exc()

finally:
    driver.quit()
